# -*- coding: utf-8 -*-
"""
File Name: init_femb.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 7/15/2016 11:47:39 AM
Last modified: 10/2/2017 12:41:28 AM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl
from openpyxl import Workbook
import numpy as np
import struct
import os
from sys import exit
import os.path
from detect_peaks import detect_peaks
from raw_convertor_m import raw_convertor
import time


def gain_process(path, wb, onedir = "step1", env = "RT", FEMB = "FEMB0", DAC = "FPGADAC" ):

    save_file = path + "\\" +onedir + "\\" + FEMB + DAC + "gain.xlsx"
    gain_data_dir = path + "\\" + onedir + "\\" 
    print gain_data_dir

    for root, dirs, files in os.walk(gain_data_dir):
        break

    for onefile in files:
        pos1 = onefile.find(FEMB)
        pos2 = onefile.find(DAC)
        #pos2 = onefile.find(DAC+"05")
        #pos2 = onefile.find("CHIP5_3F_ASICDAC07")
        pos3 = onefile.find(".bin")
        #pos3 = onefile.find("step132_FEMB3CHIP7_3D_FPGADAC04.bin")
        pos4 = onefile.find(DAC+"00")
        if (pos1 >= 0) and (pos2 >= 0) and (pos3 >= 0) and (pos4 < 0):
            gain_data_file = gain_data_dir + onefile
            print gain_data_file
            fileinfo  = os.stat(gain_data_file)
            filelength = fileinfo.st_size

            with open(gain_data_file, 'rb') as f:
                raw_data = f.read(filelength)

            smps = (filelength-1024)/2/16 

            chn_data = raw_convertor(raw_data, smps)
            chn_peakmean=[]

            for chn in range(16):
            #for chn in [4]:
                np_data = np.array(chn_data[chn])
                pedvalue = np.mean(np_data[0:10000])
                #maxvalue = np.max(np_data[0:10000])

                #print len(np_data)
                maxvalue_np = []
                for i in range(0, (len(np_data)-2000)/1000, 1):
                #    print np.max(np_data[1500*i:1500*(i+1)])
                    maxvalue_np.append( np.max(np_data[1000*i:1000*(i+1)]) )

                #print maxvalue_np
                #maxvalue_np = maxvalue_np.sort() 
                maxvalue_np =  sorted(maxvalue_np)
                #print maxvalue_np

                maxvalue = maxvalue_np[len(maxvalue_np)/5]
                peaks_index = detect_peaks(x=np_data, mph=pedvalue + abs((maxvalue-pedvalue)*2/3), mpd=500) 

                peaks_index_len = len(peaks_index)
                if ( (  peaks_index_len > 50 ) or ( peaks_index_len < 10 ) ):
                    print "ERROR, too many or too less peaks, please check!!!"
                #print maxvalue


#                peaks_cnt = len(peaks_index)
#                if (peaks_cnt < 20 ) :
#                    if ((maxvalue/2) > pedvalue ):
#                        peaks_index = detect_peaks(x=np_data, mph=pedvalue + abs(((maxvalue/2)-pedvalue)*2/3), mpd=900) 
#
#                        print len(peaks_index)
#                        for i in peaks_index :
#                            print np_data[i]
#
#                elif (peaks_cnt > 50 ) :
#                    print "ERROR, too many peaks, please check!!!"
#                    exit()
#                    


                 
                peaks_value = []
                for i in peaks_index :
                    peaks_value.append(np_data[i])

                if len(peaks_value) != 0 :
                    peaksmean = np.mean(peaks_value)
                    print "# of peaks = %d, mean =%d, delta=%d"%(len(peaks_value), np.mean(peaks_value), (np.mean(peaks_value) - pedvalue) )
                else:
                    peaksmean = pedvalue
                    print "NO Peaks"
                chn_peakmean.append(peaksmean)


            chip_pos = onefile.find("CHIP")
            chip_num = int(onefile[chip_pos+4:chip_pos+4+1],16)
            dac_pos = onefile.find(DAC)
            DAC_code = int(onefile[dac_pos+7:dac_pos+7+2],16)

            sheet_title = onefile[dac_pos-3:dac_pos-1]
            ws = wb.active
            try:
                ws = wb.get_sheet_by_name(sheet_title)
            except KeyError:
                ws = wb.create_sheet(0)
            ws.title = sheet_title #onedir[0:30]
            for chn in range(16):
                ws.cell(column=chn+1,row=chip_num+1+16*DAC_code,value=chn_peakmean[chn])             


            wb.save(filename = save_file)
    del wb
    del ws



#run01

#path = "D:\\Hibay_V2\\Rawdata\\rawdata_0930\\run04\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "stepT032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB1RTstep34"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB1RTstep32"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB2RTstep34"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB2RTstep32"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB1RTstep24"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB1RTstep22"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB2RTstep24"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0"]:
#    env = "RT"
#    onedir = "WIB2RTstep22"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

path = "D:\\Hibay_V2\\Rawdata\\rawdata_1001\\run16\\"

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB1LNstep32"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB2LNstep32"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB1LNstep34"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB2LNstep34"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB1LNstep24"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB1LNstep22"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB2LNstep24"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
    env = "LN"
    onedir = "WIB2LNstep22"
    wb = Workbook()
    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#path = "D:\\Hibay_V2\\Rawdata\\rawdata_0925\\run09\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step1112"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step1113"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step1102"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step1103"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#

#run02
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0725\\run02\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB3"]:
#    env = "RT"
#    onedir = "step033"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB3"]:
#    env = "RT"
#    onedir = "step032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB3"]:
#    env = "RT"
#    onedir = "step022"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB3"]:
#    env = "RT"
#    onedir = "step023"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )


#run03
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0725\\run03\\"
#for FEMB in ["FEMB0", "FEMB1" ,  "FEMB2"]:
#    env = "RT"
#    onedir = "step033"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" ,  "FEMB2"]:
#    env = "RT"
#    onedir = "step032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" ,  "FEMB2"]:
#    env = "RT"
#    onedir = "step022"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" ,  "FEMB2"]:
#    env = "RT"
#    onedir = "step023"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#run04
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0725\\run04\\"
#for FEMB in ["FEMB0",   "FEMB3"]:
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0",   "FEMB3"]:
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0",   "FEMB3"]:
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0",   "FEMB3"]:
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#run01
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0725\\run06\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0726\\run02\\"
#for FEMB in ["FEMB1", "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
##for FEMB in ["FEMB0"]:
#for FEMB in ["FEMB1", "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
##for FEMB in ["FEMB0"]:
#for FEMB in ["FEMB1", "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
##for FEMB in ["FEMB0"]:
#for FEMB in ["FEMB1", "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )



#####################0726
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0807\\run01\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step033"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0807\\run01\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step022"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step023"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )


#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0809\\run01\\"
#for FEMB in ["FEMB0","FEMB1",  "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0","FEMB1",  "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0","FEMB1",  "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0","FEMB1",  "FEMB2", "FEMB3"] :
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )


#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0809\\run02\\"
#for FEMB in [ "FEMB2", "FEMB3"]:
##for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step033"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step022"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step023"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )


#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0805\\run07\\"
#for FEMB in [ "FEMB3"]:
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB2", "FEMB3"]:
#for FEMB in [ "FEMB3"]:
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
##for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2"]:
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
##for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2"]:
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0726\\run04\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step113"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step112"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step102"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step103"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0804\\run04\\"
##for FEMB in ["FEMB0", "FEMB1","FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
#    env = "LN2"
#    onedir = "step133"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
##for FEMB in ["FEMB0", "FEMB1","FEMB2", "FEMB3"]:
#for FEMB in ["FEMB0"]:
#    env = "LN2"
#    onedir = "step132"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )

#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step122"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "LN2"
#    onedir = "step123"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )


#path = "D:\\FEMB_QA\\Rawdata\\rawdata_0801\\run01\\"
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step033"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step032"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step022"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "FPGADAC" )
#
#for FEMB in ["FEMB0", "FEMB1" , "FEMB2", "FEMB3"]:
#    env = "RT"
#    onedir = "step023"
#    wb = Workbook()
#    gain_process(path, wb, onedir, env, FEMB, DAC = "ASICDAC" )

