# -*- coding: utf-8 -*-
"""
File Name: read_mean.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 3/9/2016 7:12:33 PM
Last modified: Thu Oct 12 12:03:59 2017
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl

import openpyxl as px
import numpy as np
import os
from int_dac_fit import int_dac_fit
from fpga_dac_fit import fpga_dac_fit
import math
from openpyxl import Workbook

import math
import statsmodels.api as sm


def read_gain(filepath,sheetname):
    #W = px.load_workbook(filepath, use_iterators = True)
    W = px.load_workbook(filepath)
    p = W.get_sheet_by_name(name = sheetname)
    dacmean=[]
    i = 0
    for row in p.iter_rows():
        if ( (i%16)<8 ):
            for k in row:
                if (k.internal_value != None):
                    if ( math.isnan(float(k.internal_value)) ==True): 
                        dacmean.append(-1)
                    else:
                        if ( int(k.internal_value) >= 0 ):
                            #dacmean.append(k.value)
                            dacmean.append(k.internal_value)
                    #if ( int(k.internal_value) > 0 ):
                    #    dacmean.append(k.internal_value)
                else:
                    dacmean.append(-1)
        i = i + 1
    dacmean = np.resize(dacmean, [(i+8)/16, 8,16] )
    return dacmean 

def one_chn_gain(gainfile_path, chn, tp=0, env="RT",gain=3, femb=0, DAC = "FPGADAC", DACvalue = [4,5,6,7,8,9,10,11]):
    fpga_vlt_slope = fpga_dac_fit(path = "./R32_16_8_4_2_1_77iii_stength.xlsx",ideal_flg = 1)
    ln2_int_vlt_slope = int_dac_fit(1)
    rt_int_vlt_slope = int_dac_fit(0)

    if ( DAC == "FPGADAC" ):
        vlt_slope = fpga_vlt_slope
    else :
        if ( env == "RT" ):
            vlt_slope = rt_int_vlt_slope
        else:
            vlt_slope = ln2_int_vlt_slope

    #W = px.load_workbook(gainfile_path, use_iterators = True)
    W = px.load_workbook(gainfile_path)
    ws = W.get_sheet_names( )

    sheetname_list = []
    for i in ws:
        if i[0:5] != "Sheet":
            sheetname_list.append(i)

    for sheetname in sheetname_list :
        if (int(sheetname[0],16)&0x03==tp) and ((int(sheetname[1],16)>>2)&0x03==gain):
            dacmean = read_gain(gainfile_path,sheetname)

    chip = chn//16
    chipchn = chn%16
    adc_np = dacmean[:,chip,chipchn]

    dac_value = []
    adc_np_cali = []
    len_cali = len(adc_np)

    for i in range (len_cali):
        if (adc_np[i] != -1 ):
            dac_value.append(i)
            adc_np_cali.append(adc_np[i])
    dac_value = dac_value
    adc_np_cali = adc_np_cali
    x = adc_np_cali
    y = []
    for tmp in dac_value:
        y.append(vlt_slope * ( (tmp*(1.85E-13))/(1.60217646E-19) ))

    x_tmp = []
    y_tmp = []
    for onedac in DACvalue:
        x_tmp.append(x[onedac])
        y_tmp.append(y[onedac])


####    #for i in range (len_cali):
####    for i in DACvalue:
####        if (adc_np[i] != -1 ):
####            dac_value.append(i)
####            adc_np_cali.append(adc_np[i])
#####    dac_value = dac_value[0:-2]
#####    adc_np_cali = adc_np_cali[0:-2]
####    dac_value = dac_value
####    adc_np_cali = adc_np_cali
####
####    j = 0 
####    if (DAC == "ASICDAC"):
####        for j in range (len(dac_value)):
####            if (dac_value[j] >= 2):
####                break
####    x = adc_np_cali[j:]
####    y = []
####    for tmp in dac_value[j:]:
####        y.append(vlt_slope * ( (tmp*(1.85E-13))/(1.60217646E-19) ))

    

    error_fit = False 
    try:
        results = sm.OLS(y_tmp,sm.add_constant(x_tmp)).fit()
    except ValueError:
        print "Gain Error chn%d"%(chn) 
        error_fit = True 
    if ( error_fit == False ):
        error_gain = False 
        try:
            slope = results.params[1]
        except IndexError:
            slope = 0
            error_gain = True
        try:
            constant = results.params[0]
        except IndexError:
            constant = 0
    else:
        slope = 0
        constant = 0
        error_gain = True

    return chn, x, y, slope, constant, error_fit, error_gain
