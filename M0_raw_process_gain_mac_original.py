# -*- coding: utf-8 -*-
"""
File Name: init_femb.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 7/15/2016 11:47:39 AM
Last modified: Wed 06 Jun 2018 06:32:55 AM CEST
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl
from openpyxl import Workbook
import numpy as np
import struct
import os
from sys import exit
import os.path
from detect_peaks import detect_peaks
from raw_convertor_m import raw_convertor
import time


def gain_process(path, wb, onedir = "step1", env = "RT", FEMB = "FEMB0", DAC = "FPGADAC", jumbo_flag = True ):

    #save_file = path + "\\" +onedir + "\\" + FEMB + DAC + "gain.xlsx"
    save_file = path + "/" +onedir + "/" + "original" + FEMB + DAC + "gain.xlsx"
    #gain_data_dir = path + "\\" + onedir + "\\" 
    gain_data_dir = path +  onedir + "/" 
    print gain_data_dir

    for root, dirs, files in os.walk(gain_data_dir):
        break
    print files

    for onefile in files:
        pos1 = onefile.find(FEMB)
        pos2 = onefile.find(DAC)
        #pos2 = onefile.find(DAC+"05")
        #pos2 = onefile.find("CHIP5_3F_ASICDAC07")
        pos3 = onefile.find(".bin")
        #pos3 = onefile.find("step132_FEMB3CHIP7_3D_FPGADAC04.bin")
        pos4 = onefile.find(DAC+"00")
        if (pos1 >= 0) and (pos2 >= 0) and (pos3 >= 0) and (pos4 < 0):
            gain_data_file = gain_data_dir + onefile
            print gain_data_file
            fileinfo  = os.stat(gain_data_file)
            filelength = fileinfo.st_size

            with open(gain_data_file, 'rb') as f:
                raw_data = f.read(filelength)

            smps = (filelength-1024)/2/16 
            if (smps > 100000 ):
                smps = 100000
            else:
                pass

            chn_data = raw_convertor(raw_data, smps, jumbo_flag)
            chn_peakmean=[]

            for chn in range(16):
            #for chn in [4]:
                np_data = np.array(chn_data[chn])
                pedvalue = np.mean(np_data[0:10000])
                #maxvalue = np.max(np_data[0:10000])

                #print len(np_data)
                maxvalue_np = []
                for i in range(0, (len(np_data)-2000)/1000, 1):
                #    print np.max(np_data[1500*i:1500*(i+1)])
                    maxvalue_np.append( np.max(np_data[1000*i:1000*(i+1)]) )

                #print maxvalue_np
                #maxvalue_np = maxvalue_np.sort() 
                maxvalue_np =  sorted(maxvalue_np)
                #print maxvalue_np

                maxvalue = maxvalue_np[len(maxvalue_np)/5]
                peaks_index = detect_peaks(x=np_data, mph=pedvalue + abs((maxvalue-pedvalue)*2/3), mpd=300) 

                peaks_index_len = len(peaks_index)
                if ( (  peaks_index_len > 150 ) or ( peaks_index_len < 5 ) ):
                    print "ERROR, too many or too less peaks, please check!!!"
                #print maxvalue
                #exit()

                 
                peaks_value = []
                for i in peaks_index :
                    peaks_value.append(np_data[i])

                if len(peaks_value) != 0 :
                    peaksmean = np.mean(peaks_value)
                    print "# of peaks = %d, mean =%d, delta=%d"%(len(peaks_value), np.mean(peaks_value), (np.mean(peaks_value) - pedvalue) )
                else:
                    peaksmean = pedvalue
                    print "NO Peaks"
                chn_peakmean.append(peaksmean)


            chip_pos = onefile.find("CHIP")
            chip_num = int(onefile[chip_pos+4:chip_pos+4+1],16)
            dac_pos = onefile.find(DAC)
            DAC_code = int(onefile[dac_pos+7:dac_pos+7+2],16)

            sheet_title = onefile[dac_pos-3:dac_pos-1]
            ws = wb.active
            try:
                ws = wb.get_sheet_by_name(sheet_title)
            except KeyError:
                ws = wb.create_sheet(0)
            ws.title = sheet_title #onedir[0:30]
            for chn in range(16):
                ws.cell(column=chn+1,row=chip_num+1+16*DAC_code,value=chn_peakmean[chn])             
            wb.save(filename = save_file)
import sys
strdate = sys.argv[1] #
strrun = sys.argv[2]  #
strenv = sys.argv[3]

strstep = sys.argv[4]
if (strstep[-1] == "2" ):
    DAC = "FPGADAC"
elif (strstep[-1] == "4" ):
    DAC = "ASICDAC"

jumbo_flag = sys.argv[5]
if (jumbo_flag == "True"):
    jumbo_flag = True
else:
    jumbo_flag = False
server_flg = sys.argv[6]
APAno = int(sys.argv[7])


one_run =   strrun 
if (server_flg == "server"):
    path = "/nfs/rscratch/bnl_ce/shanshan/Rawdata/APA%d/Rawdata_"%APAno+ strdate + "/" + strrun + "/"
    #path = "/nfs/rscratch/bnl_ce/shanshan/Rawdata/Coldbox/Rawdata_"+ strdate + "/" + strrun + "/"
else:
    path = "/Users/shanshangao/Documents/data2/Rawdata/APA3/Rawdata_"+ strdate + "/" + strrun + "/"
#femb_set = strenv + strstep
femb_set = strstep
onedir_np = ["WIB00"+femb_set, "WIB01"+femb_set, "WIB02"+femb_set, "WIB03"+femb_set,"WIB04"+femb_set,]
for onedir in onedir_np:
    for FEMB in ["FEMB0", "FEMB1", "FEMB2","FEMB3" ]:
        Active_flg = True

        #if (APAno == 2):
        #    if (onedir.find("WIB02") >= 0) and (FEMB == "FEMB2"):
        #        Active_flg = False
        #elif (APAno == 3):
        #    if (onedir.find("WIB02") >= 0) and (FEMB == "FEMB1"):
        #        Active_flg = False
        #    elif (onedir.find("WIB02") >= 0) and (FEMB == "FEMB0"):
        #        Active_flg = False
        #    elif (onedir.find("WIB03") >= 0) and (FEMB == "FEMB0"):
        #        Active_flg = False
        #    elif (onedir.find("WIB00") >= 0) and (FEMB == "FEMB3"):
        #        Active_flg = False
        #elif (APAno == 4):
        #    if (onedir.find("WIB01") >= 0) and (FEMB == "FEMB1"):
        #        Active_flg = False

        if (Active_flg ):
            wb = Workbook()
            gain_process(path, wb, onedir, strenv, FEMB, DAC, jumbo_flag )

