# -*- coding: utf-8 -*-
"""
File Name: read_mean.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 3/9/2016 7:12:33 PM
Last modified: Fri Oct 13 22:33:54 2017
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl

import openpyxl as px
import numpy as np
import os
from int_dac_fit import int_dac_fit
from fpga_dac_fit import fpga_dac_fit
from linear_fit_excel_m import linear_fit_excel
import matplotlib.pyplot as plt
import statsmodels.api as sm
import math
from gain_plot_m import gain_plot
from openpyxl import Workbook

def read_gain(filepath,sheetname):
    #W = px.load_workbook(filepath, use_iterators = True)
    W = px.load_workbook(filepath)
    p = W.get_sheet_by_name(name = sheetname)
    
    dacmean=[]
    i = 0
    for row in p.iter_rows():
        if ( (i%16)<8 ):
            for k in row:
                #if (k.value != None):
                if (k.internal_value != None):
                    if ( math.isnan(float(k.internal_value)) ==True): 
                        dacmean.append(-1)
                    else:
                        if ( int(k.internal_value) >= 0 ):
                            #dacmean.append(k.value)
                            dacmean.append(k.internal_value)
                else:
                    dacmean.append(-1)
        i = i + 1
    dacmean = np.resize(dacmean, [(i+8)/16, 8,16] )
    return dacmean 

def get_gain_results(root, filename, env = "RT", DAC = "FPGADAC" ):
    #filepath= root + "\\" + filename
    filepath= root + "/" + filename

    #W = px.load_workbook(filepath, use_iterators = True)
    W = px.load_workbook(filepath)
    ws = W.get_sheet_names( )

    sheetname_list = []
    for i in ws:
        if i[0:5] != "Sheet":
            sheetname_list.append(i)

    fpga_vlt_slope = fpga_dac_fit(path = "./R32_16_8_4_2_1_77iii_stength.xlsx",ideal_flg = 1)
    ln2_int_vlt_slope = int_dac_fit(1)
    rt_int_vlt_slope = int_dac_fit(0)

    if ( DAC == "FPGADAC" ):
        vlt_slope = fpga_vlt_slope
    else :
        if ( env == "RT" ):
            vlt_slope = rt_int_vlt_slope
        else:
            vlt_slope = ln2_int_vlt_slope

    gain_result = []
    for sheetname in sheetname_list :
        #result_path= root + "\\" + filename[0:12] + "\\" + sheetname + "_fits\\" 
        result_path= root + "/" + filename[0:20] + "/" + sheetname + "_fits/" 
        try: 
            os.makedirs(result_path)
        except OSError:
            if os.path.exists(result_path):
                pass
    
        #plot_path= result_path + "\\" 
        plot_path= result_path + "/" 
        try: 
            os.makedirs(plot_path)
        except OSError:
            if os.path.exists(plot_path):
                pass
    
        dacmean = read_gain(filepath,sheetname)

        slope = []
        constant = []
        for chip_id in range(8):
            plt.figure(figsize=(16,9))
            x_plot = np.linspace(0,4100)
    
            plt.ylabel("electrons")
            plt.xlabel("ADC counts")
            plt.title("Calibration of chip%d"%(chip_id+1) )

            for chn in range(16):
#            for chn in [5]:
                adc_np = dacmean[:,chip_id,chn]
                dac_value = []
                adc_np_cali = []
                len_cali = len(adc_np)
                for i in range (len_cali):
                    if (adc_np[i] != -1 ):
                        dac_value.append(i)
                        adc_np_cali.append(adc_np[i])

#                dac_value = dac_value[1:11]
#                adc_np_cali = adc_np_cali[1:11]

#                dac_value = dac_value[0:-2]
#                adc_np_cali = adc_np_cali[0:-2]
#                dac_value = [dac_value[0],dac_value[2],dac_value[4],dac_value[6]]
#                adc_np_cali = [adc_np_cali[0],adc_np_cali[2],adc_np_cali[4],adc_np_cali[6]]

                #tmp_len = len(adc_np_cali)
                #for i in range(tmp_len-3):
                #    if ( (adc_np_cali[i+2] > adc_np_cali[i+3]) ) or ( adc_np_cali[i+3] > 4100 ):
                #        break
                #dac_value = dac_value[0:i+3+1]
                #adc_np_cali = adc_np_cali[0:i+3+1]
                j = 0 
                if (DAC == "ASICDAC"):
                    for j in range (len(dac_value)):
                        if (dac_value[j] >= 2):
                            break

                x = adc_np_cali[j:]
                y = []
#                if len(x) != 4:
#                    print len(x)
                for tmp in dac_value[j:]:
                    y.append(vlt_slope * ( (tmp*(1.85E-13))/(1.60217646E-19) ))

                x_min = 0
                x_max = 4 
                a, b= linear_fit_excel(chip_id,chn, x, y, x_min, x_max, plot_path, plot_en = 0)
                slope.append(a)
                constant.append(b)

        gain_result.append([sheetname,slope])
        gain_plot(sheetname, slope, save_path = root , env=env , FEMB = filename[0:13],  DAC=DAC )
        
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for gain in gain_result:
        ws = wb.create_sheet(0)
        ws.title = gain[0]
        for i in range(len(gain[1])):
            ws.cell(column=1,row=i+1,value=gain[1][i])
    
    #wb.save(filename = root + "\\" + "Calcuated_" + filename[0:-5]+".xlsx")
    wb.save(filename = root + "/" + "Calcuated_" + filename[0:-5]+".xlsx")
    del wb
    del ws



env = "RT"
#

for FEMB in ["FEMB0","FEMB1", "FEMB2","FEMB3"]:
#for FEMB in ["FEMB0"]:
    onedir = "WIB5RTstep32"
    DAC = "FPGADAC"
    #DAC = "ASICDAC"

    root_path ="/Users/shanshangao/Documents/Share_Windows/CERN_test_stand/Rawdata/Rawdata_1013/run01/" + onedir   
    filename = "original" + FEMB + DAC + "gain.xlsx"
    get_gain_results(root_path,filename)

